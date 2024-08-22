from tkinter import messagebox, filedialog
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font
import datetime
import os

userDesktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')

file = filedialog.askopenfilename(title="Select the file to convert", initialdir=userDesktop, filetypes=[("Excel files", "*.xlsx")])

if not file:
    os._exit(0)
    
wb = load_workbook(file)
ws = wb.active

# "Pay Period: 8/19/2024 - 8/19/2024"

combDate = ws['A1'].value
startDate = combDate.split(': ')[1].split(' - ')[0]
endDate = combDate.split(': ')[1].split(' - ')[1]

# Create a new csv file

username = os.getlogin()
outputFolder = f"C:\\Users\\{username}\\Desktop\\"
outputFile = outputFolder + 'ADP_Upload.csv'

try:

    with open(outputFile, 'w') as f:
        

        data = []
        for row in ws.iter_rows():
            # skip first row
            # add columns A, C:I to data
            
            if row[0].row == 1 or row[0].row == ws.max_row:
                continue
            
            for cell in row:
                
                # if cell.column in ['C', 'D', 'E', 'F', 'G', 'H', 'I']:
                if cell.column in [3, 4, 5, 6, 7, 8, 9] and cell.value and cell.value != '0':
                    #Regular Hours	Overtime Hours	Double Time Hours	Vacation Hours	Sick Hours	Holiday Hours	Personal Hours

                    workTypes = {
                        # 'C' : 'REG',
                        # 'D' : 'OT',
                        # 'E' : 'DT',
                        # 'F' : 'VAC',
                        # 'G' : 'SICK',
                        # 'H' : 'HOL',
                        # 'I' : 'PER',
                        3 : 'REG',
                        4 : 'OT',
                        5 : 'DT',
                        6 : 'VAC',
                        7 : 'SICK',
                        8 : 'HOL',
                        9 : 'PER',
                        
                    }
                    
                    empID = ws['A' + str(cell.row)].value
                    if str(empID).isnumeric():
                        data.append([ws['A' + str(cell.row)].value, cell.value, workTypes[cell.column]])
                    

        # Write data to new workbook
        # header: Company Code	 Pay Frequency	 Start Date	 End Date	 Employee ID	 Earnings Code	 Pay Hours	 Separate Check	 Rate Code
        # data: B	B	8/1/2024	8/19/2024	6	REG	95.5	0	BASE


        f.write('Company Code, Pay Frequency, Start Date, End Date, Employee ID, Earnings Code, Pay Hours, Separate Check, Rate Code\n')

        for row in data:
            f.write(f'B,B,{startDate},{endDate},{str(row[0])},{row[2]},{str(row[1])},0,BASE\n')

except PermissionError:
    messagebox.showerror("Error", "The ADP_Upload.csv file is open. Please close the file and try again.")
    os._exit(0)
except Exception as e:
    messagebox.showerror("Error", str(e))
    os._exit(0)

messagebox.showinfo("Success", "The new file has been saved")
