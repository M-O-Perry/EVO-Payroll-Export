from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font
import datetime
import os

class ReportOutput:
    def __init__(self, employeeTimeSheet, shiftSummary, start_date, end_date):
        username = os.getlogin()
        self.outputFolder = f"C:\\Users\\{username}\\Desktop\\"
        #self.outputFolder = "\\\\FS2\\engineer\\WORK\\Outdwg\\SolidworksBomOutputs\\"
        
        self.start_date = start_date
        self.end_date = end_date
        self.employeeTimeSheet = employeeTimeSheet
        self.shiftSummary = shiftSummary
        self.EVO_file_path = self.outputFolder + "EVO_Payroll.xlsx"
        self.ADP_file_path = self.outputFolder + "ADPUpload.csv"
    
    def print_all(self):
        ss = self.print_shift_summary()
        ts = self.print_time_sheet()

        return "Shift summary:\n" + ss + "\nTime sheet summary:\n" + ts
        
    
    def print_shift_summary(self):
        output = ""
        
        for key, value in self.shiftSummary.items():
            output += f"{key}:\n"
            for item in value:
                if isinstance(item, dict):
                    for date, hours in item.items():
                        output += f"        {date}: {hours}\n"
                else:
                    output += f"    {item},\n"
        
        print(output)
        return output
    
    def print_time_sheet(self):
        output = ""
        
        for emp, value in self.employeeTimeSheet.items():
            output += f"{emp}:\n"
            
            for date, entries in value.items():
                output += f"    {date}: {entries}\n"
                
        
        print(output)
        return output

    def write_to_excel(self):
        wb = Workbook()
        ws = wb.active
        
        ws.append(["Employee ID", "Name", "Regular Hours", "Overtime Hours", "Double Time Hours", "Vacation Hours", "Sick Hours", "Holiday Hours", "Personal Hours", "Total Hours"])
        
        totalTimes = [0] * len(self.shiftSummary[list(self.shiftSummary.keys())[0]][1:-1])
        
        for emp in self.shiftSummary:
            hours = self.shiftSummary[emp]
            
            #print(hours)
            
            output = [emp]
        
            output.append(hours[0]) # Name
            
            for i in range(1,len(hours[1:-1])+1):
                #print(i, hours[i])
                if i != 1 and hours[i] == 0:  # hours[1] = Regular Hours, want to also include 0 hours unlike other categories
                    output.append("")
                else:
                    output.append(hours[i])
                    
                totalTimes[i-1] += float(hours[i])
            
            output.append(sum(hours[1:-1]))
            
            ws.append(output)
            
        ws.append([])
        ws.append(["Total:", "", *totalTimes, sum(totalTimes)])
        
            
        
        wb.create_sheet('Detailed Comparison')
        ws = wb['Detailed Comparison']
        
        ws.append(["Employee ID", "Date", "Total Clock Time", "Total Shift Time"])
        
        combinedEmployees = {}
        
        for emp in self.employeeTimeSheet:
            combinedEmployees[emp] = self.employeeTimeSheet[emp]["name"]
        
        for emp in self.shiftSummary:
            combinedEmployees[emp] = self.shiftSummary[emp][0]
            
        
        for emp in list(sorted(combinedEmployees.keys(), key = lambda x: int(x))):
            name = combinedEmployees[emp]
                
            ws.append([emp, name])
            
            allEntryDates = set()
            
            if emp in self.employeeTimeSheet:
                for date, hours in self.employeeTimeSheet[emp].items():
                    allEntryDates.add(date)
            
            if emp in self.shiftSummary:
                for date, hours in self.shiftSummary[emp][-1].items():
                    allEntryDates.add(date)
                
            allEntryDates = sorted(list(allEntryDates))
            #print(allEntryDates)
            if "name" in allEntryDates:
                allEntryDates.remove("name")
            
            #print("Entry Dates:   ", allEntryDates)
            
            for date in allEntryDates:
                
                shiftHours = 0
                clockHours = 0
                    
                if emp in self.employeeTimeSheet and date in self.employeeTimeSheet[emp]:
                    clockHours = float(self.employeeTimeSheet[emp][date])
                
                if emp in self.shiftSummary and date in self.shiftSummary[emp][-1]:
                    shiftHours = float(self.shiftSummary[emp][-1][date])     
                
                output = [""]
                
                output.append(date)
                output.append(clockHours)
                output.append(shiftHours)
                
                overReportedEpsilon = 0.25
                underReportedEpsilon = -2
                
                if (shiftHours - clockHours) > overReportedEpsilon:
                    output.append("")
                    output.append("MISMATCH")
                
                elif (shiftHours - clockHours) < underReportedEpsilon:
                    output.append("")
                    output.append("Under-reported")
                
                ws.append(output)
                
            ws.append([])
            
            
        try:
            wb.save(self.EVO_file_path)
        except PermissionError:
            # tkinter popup message to allert user that the file is open
            
            messagebox.showinfo("Error", "The output file (EVO_Payroll.xlsx) is open. Please close the file and re-run the program again.")
            
        wb.close()

    def output_to_ADP(self, employeeShifts):
        self.createADPFile()
        
            
    
    def createADPFile(self):
        # Create a csv file with the ADP format: Company Code, Pay Frequency, Start Date, End Date, Employee ID, Earnings Code, Pay Hours, Separate Check, Rate Code
        adp_format = "Company Code, Pay Frequency, Start Date, End Date, Employee ID, Earnings Code, Pay Hours, Separate Check, Rate Code"
        
        try:
            with open(self.ADP_file_path, "w") as file:
                file.write(adp_format)
        except PermissionError:
            # tkinter popup message to allert user that the file is open
            
            messagebox.showinfo("Error", "The output file (ADPUpload.csv) is open. Please close the file and re-run the program again.")
        
    def addAllEmployeesToADP(self):
        employees = self.shiftSummary.keys()
        
        with open(self.ADP_file_path, "a") as file:
            file.write("\n")
            for emp in employees:
                #print("new Employee: ", emp)
                
                company_code = "B"
                pay_frequency = "B"
                start_date = self.__formatDate(self.start_date)
                end_date = self.__formatDate(self.end_date)
                empID = emp
                separate_check = 0
                rate_code = "BASE"
                
                type = ""
                hours = ""
                
                hourTypes = ["REG", "OT", "DT", "VAC", "SICK", "HOL", "PER"]
                
                #print(self.shiftSummary[emp])
                
                for i in range(1, len(self.shiftSummary[emp][1:-1])+1):
                    shift = self.shiftSummary[emp][i]
                    
                    print(shift)
                    if shift:
                        hours = shift
                        type = hourTypes[i-1]
                
                        file.write(f"{company_code},{pay_frequency},{start_date},{end_date},{empID},{type},{hours},{separate_check},{rate_code}\n")
                        #print(f"{company_code},{pay_frequency},{start_date},{end_date},{empID},{type},{hours},{separate_check},{rate_code}")
        
    def __formatDate(self, date):
        # date format is MMDDYY
        # should be MM/DD/YYYY
        
        year = "19"
        
        if date[4:] < str(int(datetime.datetime.now().strftime("%y")) + 1):
            year = "20" + date[4:]
        
        #print(f"{date[:2]}/{date[2:4]}/{year}")
        return f"{date[:2]}/{date[2:4]}/{year}"
            
            
            
            
    def formatExcelOutput(self):
        wb = load_workbook(self.EVO_file_path)
        ws = wb.active
        
        self.__formatSheet(ws)
        
        ws = wb['Detailed Comparison']
        self.__formatSheet(ws)
        
        
        wb.save(self.EVO_file_path)
        
    def __formatSheet(self, ws):
        thin = Side(style = 'thin')
        medium = Side(style = 'medium')
        thick = Side(style = 'thick')
        
        dims = {}
        
        for row in ws.iter_rows():
            for cell in row:
                top, left, right, bottom = thin, thin, thin, thin
                
                if cell.row % 2 == 1 and cell.row != ws.max_row-1:
                    cell.fill = PatternFill(start_color = "e6e6e6", end_color = "e6e6e6", fill_type = "solid")

                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
                    
                    
                if cell.row == 1:
                    top = thick
                    bottom = medium
                    cell.font = Font(size=12)
                    
                if cell.row == ws.max_row:
                    bottom = thick
                    
                if cell.column == 1:
                    left = thick
                    cell.font = Font(size=12)
                    
                    if cell.row != ws.max_row and cell.row != 1:          
                        right = medium
                        cell.font = Font(size = 12)
                        
                if cell.column == ws.max_column:
                    right = thick
                    
                    if cell.value == "MISMATCH":
                        cell.fill = PatternFill(start_color = "ffcccb", end_color = "ffcccb", fill_type = "solid")
                        cell.font = Font(bold = True)
                    elif cell.value == "Under-reported": 
                        cell.fill = PatternFill(start_color = "ffff99", end_color = "ffff99", fill_type = "solid")
                
                
                cell.border = Border(top = top, left = left, right = right, bottom = bottom)
                
        
        for col, value in dims.items():
            ws.column_dimensions[chr(64+col)].width = value + 2